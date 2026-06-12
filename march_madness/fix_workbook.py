"""Write March_Madness_20190318_fixed.xlsx with the Value-pick pipeline repaired.

Two changes to the 'Slot calculations' tab (see extract_model.py for the full
story and verification):

1. Variance difference, AL3:AQ66 -- the S term referenced the cell 38 rows
   below its own row (AL3 said B41; from row 29 the reference runs off the
   data entirely). The paper's formula uses the slot's own S:
   dVar = (p - p0)(1 + 2S) - p^2 + p0^2 - 2(p*f - p0*f0).

2. Value flag, BD3:BI66 -- was IF(OR(ratio=0, ratio>1),1,0). With the sheet's
   sign convention (cumulative expected-point difference <= 0 for
   non-favorites) the paper's rule -- variance gain at least as large as the
   expected-value loss -- is ratio <= -1, so the test becomes
   IF(OR(ratio=0, ratio<=-1),1,0).

Everything else is preserved as-is. Cell values are recalculated when the file
is next opened in Excel/LibreOffice.
"""

from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
SOURCE = HERE / "reference" / "March_Madness_20190318.xlsx"
DEST = HERE / "March_Madness_20190318_fixed.xlsx"

# (variance-diff col, prob/freq col, p0 col, p0f0 col, Bracket points col)
VAR_COLS = [("AL", "B", "H", "N", "E"), ("AM", "C", "I", "O", "F"),
            ("AN", "D", "J", "P", "G"), ("AO", "E", "K", "Q", "H"),
            ("AP", "F", "L", "R", "I"), ("AQ", "G", "M", "S", "J")]
# (value flag col, value ratio col)
FLAG_COLS = [("BD", "AX"), ("BE", "AY"), ("BF", "AZ"),
             ("BG", "BA"), ("BH", "BB"), ("BI", "BC")]


def main():
    wb = openpyxl.load_workbook(SOURCE, data_only=False)
    ws = wb["Slot calculations"]
    for r in range(3, 67):
        pf = r - 1  # 'Brackets - prob'/'freq' rows are offset by one
        for vc, c, p0c, pfc, ptc in VAR_COLS:
            ws[f"{vc}{r}"] = (
                f"=Bracket!{ptc}$2^2*(('Brackets - prob'!{c}{pf}-{p0c}{r})"
                f"*(1+2*{c}{r})-'Brackets - prob'!{c}{pf}^2+{p0c}{r}^2"
                f"-2*('Brackets - prob'!{c}{pf}*'Brackets - freq'!{c}{pf}-{pfc}{r}))")
        for fc, rc in FLAG_COLS:
            ws[f"{fc}{r}"] = f"=IF(OR({rc}{r}=0,{rc}{r}<=-1),1,0)"
    wb.save(DEST)
    print(f"Wrote {DEST}")


if __name__ == "__main__":
    main()
