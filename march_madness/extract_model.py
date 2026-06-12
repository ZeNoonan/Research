"""Extract Aaron Brown's March Madness 2019 factor model from the source xlsx.

Reads reference/March_Madness_20190318.xlsx and writes data/model_2019.json with
everything the webpage needs: teams, seeds, per-round win probabilities and pick
frequencies, the Quality/Size/Momentum factor values and the default picks.

It also re-implements the 'Slot calculations' tab in Python in two modes:

published -- reproduces the workbook exactly as released, including two bugs in
  the Value-pick pipeline:
  (1) the 'Variance difference' columns AL:AQ read the slot S term from 38 rows
      below their own row (AL3 says B41, a fill-down error; from row 29 the
      reference runs off the data and Excel treats the empty cells as 0);
  (2) the Value flag tests ratio > 1, but with the sheet's sign convention
      (cumulative expected-point difference is <= 0) a genuine value pick --
      variance gain at least as large as the expected-value loss, the paper's
      stated rule -- has ratio <= -1.  As published the flag fires for picks
      whose variance LOSS exceeds the EV loss, the opposite of the paper.

fixed -- same-row S in the variance difference, per the paper's formula
  dVar = (p-p0)(1+2S) - p^2 + p0^2 - 2(p*f - p0*f0), and the Value flag set
  when the pick is the favorite (ratio 0) or ratio <= -1.

Run with --report to verify the published mode cell-by-cell against the
workbook's cached values and to see the impact of the fixes.
"""

from __future__ import annotations

import json
import math
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
SOURCE = HERE / "reference" / "March_Madness_20190318.xlsx"
OUT = HERE / "data" / "model_2019.json"

ROUNDS = ["R64", "R32", "S16", "E8", "F4", "C6"]
REGIONS = ["East", "West", "South", "Midwest"]
N_TEAMS = 64
# AL3 says B41: 38 rows below its own row 3. Filled down, every row points 38
# rows past itself; rows 29-66 land on empty cells which Excel treats as 0.
BUG_OFFSET = 38


def excel_lt(a, b):
    """Excel-style a < b: operands equal to 15 significant digits compare equal."""
    if a == b:
        return False
    scale = max(abs(a), abs(b))
    if scale and abs(a - b) / scale < 1e-14:
        return False
    return a < b


def load():
    wb_f = openpyxl.load_workbook(SOURCE, data_only=False)
    wb_v = openpyxl.load_workbook(SOURCE, data_only=True)
    return wb_f, wb_v


def extract(wb_v):
    """Pull raw model inputs from the workbook (cached values only)."""
    prob_ws = wb_v["Brackets - prob"]
    freq_ws = wb_v["Brackets - freq"]
    bracket = wb_v["Bracket"]

    teams = []
    for i in range(N_TEAMS):
        r_pf = i + 2   # prob/freq sheets: rows 2-65
        r_br = i + 14  # bracket tab: rows 14-77
        teams.append({
            "seed": bracket.cell(row=r_br, column=1).value,
            "name": bracket.cell(row=r_br, column=2).value,
            "longName": prob_ws.cell(row=r_pf, column=1).value,
            "region": REGIONS[i // 16],
            "prob": [prob_ws.cell(row=r_pf, column=2 + k).value for k in range(6)],
            "freq": [freq_ws.cell(row=r_pf, column=2 + k).value for k in range(6)],
            # Q/R/S columns: Quality, Size, Momentum (play-in rows are the
            # average of the two candidate teams, so take the cached value)
            "quality": bracket.cell(row=r_br, column=17).value,
            "size": bracket.cell(row=r_br, column=18).value,
            "mom": bracket.cell(row=r_br, column=19).value,
            "picks": [bracket.cell(row=r_br, column=5 + k).value for k in range(6)],
        })

    # Known data quirk in the source workbook: the South 8 seed (2019: Ole
    # Miss) is listed as 'Mississippi State' -- already the East 5 seed -- so
    # its Quality/Size/Momentum lookups return Mississippi State's values.
    south8 = teams[34]
    if south8["name"] == "Mississippi State" and south8["seed"] == 8:
        south8["note"] = ("Listed as 'Mississippi State' in the source workbook; "
                          "the 2019 South 8 seed was Mississippi (Ole Miss). Its "
                          "Quality/Size/Momentum values are Mississippi State's.")

    points = [bracket.cell(row=2, column=5 + k).value for k in range(6)]
    pool_size = bracket.cell(row=3, column=4).value
    return {"year": 2019, "rounds": ROUNDS, "points": points,
            "poolSize": pool_size, "teams": teams}


def slot_calcs(model, mode):
    """Re-implementation of the 'Slot calculations' tab ('published'/'fixed')."""
    assert mode in ("published", "fixed")
    teams, points = model["teams"], model["points"]
    out = []
    # Slot groups: round k joins blocks of 2^(k+1) consecutive teams
    groups = []
    for k in range(6):
        size = 2 ** (k + 1)
        groups.append([range(g, g + size) for g in range(0, N_TEAMS, size)])

    S = [[0.0] * 6 for _ in range(N_TEAMS)]
    p0 = [[0.0] * 6 for _ in range(N_TEAMS)]
    p0f0 = [[0.0] * 6 for _ in range(N_TEAMS)]
    for k in range(6):
        for grp in groups[k]:
            s = sum(teams[i]["prob"][k] * teams[i]["freq"][k] for i in grp)
            mx = max(teams[i]["prob"][k] for i in grp)
            tied = [teams[i]["freq"][k] for i in grp if teams[i]["prob"][k] == mx]
            pf = (sum(tied) / len(tied)) * mx
            for i in grp:
                S[i][k], p0[i][k], p0f0[i][k] = s, mx, pf

    for i, t in enumerate(teams):
        exp_diff, var_one, var_diff = [], [], []
        for k in range(6):
            p, f = t["prob"][k], t["freq"][k]
            exp_diff.append(points[k] * (p - p0[i][k]))
            s = S[i][k]
            # AF:AK 'Variance': team's own contribution to score variance
            var_one.append(p * (1 - p - 2 * f + 2 * s) + s - s * s)
            if mode == "fixed":
                s_used = s
            else:
                j = i + BUG_OFFSET
                s_used = S[j][k] if j < N_TEAMS else 0.0
            var_diff.append(points[k] ** 2 * (
                (p - p0[i][k]) * (1 + 2 * s_used)
                - p * p + p0[i][k] ** 2
                - 2 * (p * f - p0f0[i][k])))

        cum_exp = [sum(exp_diff[: k + 1]) for k in range(6)]
        cum_var = [sum(var_diff[: k + 1]) for k in range(6)]
        ratio = [0.0 if cum_exp[k] == 0 else cum_var[k] / cum_exp[k] for k in range(6)]
        if mode == "published":
            value = [1 if (ratio[k] == 0 or ratio[k] > 1) else 0 for k in range(6)]
        else:
            value = [1 if (ratio[k] == 0 or ratio[k] <= -1) else 0 for k in range(6)]
        bab = [1 if (excel_lt(p0[i][k] - 0.5, t["prob"][k])
                     and excel_lt(t["freq"][k] / t["prob"][k],
                                  p0f0[i][k] / p0[i][k] ** 2))
               else 0 for k in range(6)]
        out.append({"S": S[i], "p0": p0[i], "p0f0": p0f0[i],
                    "expDiff": exp_diff, "cumExp": cum_exp, "variance": var_one,
                    "varDiff": var_diff, "cumVar": cum_var, "ratio": ratio,
                    "value": value, "bab": bab})
    return out


def factor_indicators(model):
    """Bracket tab V:X -- Quality/Size above seed median, Momentum rule."""
    teams = model["teams"]
    med_q, med_s = {}, {}
    for seed in range(1, 17):
        qs = sorted(t["quality"] for t in teams if t["seed"] == seed)
        ss = sorted(t["size"] for t in teams if t["seed"] == seed)
        med_q[seed] = (qs[1] + qs[2]) / 2
        med_s[seed] = (ss[1] + ss[2]) / 2
    ind = []
    for t in teams:
        ind.append({
            "quality": 1 if t["quality"] > med_q[t["seed"]] else 0,
            "size": 1 if t["size"] > med_s[t["seed"]] else 0,
            "mom": 1 if ((t["seed"] < 9 and t["mom"] > 3)
                         or (t["seed"] > 8 and t["mom"] < 4)) else 0,
        })
    return ind, med_q, med_s


def stars(model, slots):
    """Bracket tab K:P -- star count per team per round."""
    ind, _, _ = factor_indicators(model)
    return [[ind[i]["quality"] + ind[i]["size"] + ind[i]["mom"]
             + slots[i]["value"][k] + slots[i]["bab"][k]
             for k in range(6)] for i in range(N_TEAMS)]


def norm_cdf(x, mu, sigma):
    return 0.5 * (1 + math.erf((x - mu) / (sigma * math.sqrt(2))))


def norm_inv(p):
    lo, hi = -40.0, 40.0
    for _ in range(200):
        x = (lo + hi) / 2
        if norm_cdf(x, 0, 1) < p:
            lo = x
        else:
            hi = x
    return (lo + hi) / 2


def win_probability(model, slots, picks):
    """Bracket tab top-right block: your chance of winning the pool."""
    teams, points, n = model["teams"], model["points"], model["poolSize"]
    you_mean = sum(points[k] * picks[i][k] * teams[i]["prob"][k]
                   for i in range(N_TEAMS) for k in range(6))
    field_mean = sum(points[k] * teams[i]["prob"][k] * teams[i]["freq"][k]
                     for i in range(N_TEAMS) for k in range(6))
    you_var = sum(points[k] ** 2 * picks[i][k] * slots[i]["variance"][k]
                  for i in range(N_TEAMS) for k in range(6))
    field_var = sum(points[k] ** 2 * teams[i]["freq"][k] * slots[i]["variance"][k]
                    for i in range(N_TEAMS) for k in range(6))
    d10, e10, e11 = you_mean - field_mean, math.sqrt(you_var), math.sqrt(field_var)
    total_pts = 32 * points[0] + 16 * points[1] + 8 * points[2] \
        + 4 * points[3] + 2 * points[4] + points[5]

    qs = [1.0 / n]
    for _ in range(6):
        qs.append(min(1.0, 2 * qs[-1]))
    qs = qs[::-1]                      # O4..V4
    for _ in range(6):
        qs.append(qs[-1] / 2)          # W4..AA4
    qs.append(0.0)                     # AB4
    xs = [-total_pts] + [d10 + norm_inv(1 - q) * e10 for q in qs[1:-1]] + [total_pts]
    Fs = [0.0] + [norm_cdf(x, 0, e11) ** (n - 1) for x in xs[1:-1]] + [1.0]
    win = sum((qs[j - 1] - qs[j]) * (Fs[j - 1] + Fs[j]) / 2 for j in range(1, 14))
    return {"youMean": you_mean, "fieldMean": field_mean, "youVar": you_var,
            "fieldVar": field_var, "winProb": win}


def verify(wb_v, model, slots):
    """Compare the published-mode re-implementation against cached values."""
    ws = wb_v["Slot calculations"]
    cols = {"S": 2, "p0": 8, "p0f0": 14, "expDiff": 20, "cumExp": 26,
            "variance": 32, "varDiff": 38, "cumVar": 44, "ratio": 50,
            "value": 56, "bab": 62}
    worst, n_checked = 0.0, 0
    for i in range(N_TEAMS):
        for key, c0 in cols.items():
            for k in range(6):
                cached = ws.cell(row=i + 3, column=c0 + k).value
                if cached is None:
                    continue
                worst = max(worst, abs(cached - slots[i][key][k]))
                n_checked += 1
    print(f"Slot calculations: {n_checked} cells checked, "
          f"max abs diff vs workbook = {worst:.3e}")

    br = wb_v["Bracket"]
    star_counts = stars(model, slots)
    mismatches = sum(
        1 for i in range(N_TEAMS) for k in range(6)
        if len(br.cell(row=i + 14, column=11 + k).value or "") != star_counts[i][k])
    print(f"Bracket star ratings: {N_TEAMS * 6} cells checked, "
          f"{mismatches} mismatches vs workbook")

    picks = [t["picks"] for t in model["teams"]]
    wp = win_probability(model, slots, picks)
    for label, cell, ours in [("You expected points", "D5", wp["youMean"]),
                              ("Field expected points", "D6", wp["fieldMean"]),
                              ("You variance", "D7", wp["youVar"]),
                              ("Field variance", "D8", wp["fieldVar"]),
                              ("Win probability", "O1", wp["winProb"])]:
        cached = br[cell].value
        print(f"{label}: ours={ours:.10f} cached={cached:.10f} "
              f"diff={abs(ours - cached):.3e}")


def report_fix_impact(model, published, fixed):
    teams = model["teams"]
    changes = [(teams[i], ROUNDS[k], published[i]["value"][k], fixed[i]["value"][k])
               for i in range(N_TEAMS) for k in range(6)
               if published[i]["value"][k] != fixed[i]["value"][k]]
    gained = [(t, r) for t, r, p, f in changes if f == 1]
    lost = [(t, r) for t, r, p, f in changes if f == 0]
    print(f"\nValue flags changed by the fixes: {len(changes)} "
          f"({len(gained)} gained, {len(lost)} lost)")
    print("Gained:")
    for t, r in gained:
        print(f"  {t['name']} ({t['seed']} seed) {r}")
    print("Lost:")
    for t, r in lost:
        print(f"  {t['name']} ({t['seed']} seed) {r}")


def main():
    wb_f, wb_v = load()
    model = extract(wb_v)
    published = slot_calcs(model, "published")
    fixed = slot_calcs(model, "fixed")

    if "--report" in sys.argv:
        verify(wb_v, model, published)
        report_fix_impact(model, published, fixed)

    OUT.parent.mkdir(exist_ok=True)
    OUT.write_text(json.dumps(model, indent=1))
    print(f"\nWrote {OUT}")


if __name__ == "__main__":
    main()
